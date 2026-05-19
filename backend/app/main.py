from datetime import datetime, timedelta, timezone
from decimal import Decimal
from enum import Enum
from typing import List, Optional
import re
import unicodedata

from fastapi import Depends, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from jose import jwt
from pydantic import BaseModel
from sqlalchemy import DateTime, Enum as SAEnum, ForeignKey, Numeric, String, Text, create_engine
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column, relationship, sessionmaker
from openpyxl import load_workbook

SECRET = "change-me-in-production"
ALGO = "HS256"


class Base(DeclarativeBase):
    pass


class Role(str, Enum):
    admin = "admin"
    sales = "sales"


class SupplierLine(str, Enum):
    acronis = "acronis"
    fortinet = "fortinet"
    bare_metal = "bare_metal"
    colocation = "colocation"


class User(Base):
    __tablename__ = "users"
    id: Mapped[int] = mapped_column(primary_key=True)
    username: Mapped[str] = mapped_column(String(60), unique=True)
    password: Mapped[str] = mapped_column(String(120))
    role: Mapped[Role] = mapped_column(SAEnum(Role))


class Setting(Base):
    __tablename__ = "settings"
    id: Mapped[int] = mapped_column(primary_key=True)
    line: Mapped[SupplierLine] = mapped_column(SAEnum(SupplierLine), index=True)
    tax_percent: Mapped[Decimal] = mapped_column(Numeric(10, 4), default=Decimal("14.53"))
    markup_percent: Mapped[Decimal] = mapped_column(Numeric(10, 4), default=Decimal("20"))
    hour_cost: Mapped[Decimal] = mapped_column(Numeric(10, 4), default=Decimal("24"))


class PriceListVersion(Base):
    __tablename__ = "price_list_versions"
    id: Mapped[int] = mapped_column(primary_key=True)
    line: Mapped[SupplierLine] = mapped_column(SAEnum(SupplierLine), index=True)
    supplier_name: Mapped[str] = mapped_column(String(120))
    uploaded_by: Mapped[str] = mapped_column(String(60))
    created_at: Mapped[datetime] = mapped_column(DateTime, default=lambda: datetime.now(timezone.utc))
    is_active: Mapped[bool] = mapped_column(default=True)
    items: Mapped[List["PriceItem"]] = relationship(back_populates="version", cascade="all,delete")


class PriceItem(Base):
    __tablename__ = "price_items"
    id: Mapped[int] = mapped_column(primary_key=True)
    version_id: Mapped[int] = mapped_column(ForeignKey("price_list_versions.id"), index=True)
    sku: Mapped[str] = mapped_column(String(120), index=True)
    description: Mapped[str] = mapped_column(Text)
    unit_cost: Mapped[Decimal] = mapped_column(Numeric(14, 6))
    currency: Mapped[str] = mapped_column(String(10), default="BRL")
    type: Mapped[str] = mapped_column(String(30), default="default")
    version: Mapped[PriceListVersion] = relationship(back_populates="items")


class Quote(Base):
    __tablename__ = "quotes"
    id: Mapped[int] = mapped_column(primary_key=True)
    line: Mapped[SupplierLine] = mapped_column(SAEnum(SupplierLine), index=True)
    created_by: Mapped[str] = mapped_column(String(60))
    customer_name: Mapped[Optional[str]] = mapped_column(String(120), nullable=True)
    payload: Mapped[str] = mapped_column(Text)
    total_monthly: Mapped[Decimal] = mapped_column(Numeric(14, 2))
    total_contract: Mapped[Decimal] = mapped_column(Numeric(14, 2))
    created_at: Mapped[datetime] = mapped_column(DateTime, default=lambda: datetime.now(timezone.utc))


engine = create_engine("sqlite:////data/pricing.db", connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(bind=engine)
Base.metadata.create_all(engine)

app = FastAPI(title="GOX Pricing Platform")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

security = HTTPBearer()


class LoginIn(BaseModel):
    username: str
    password: str


class TokenOut(BaseModel):
    access_token: str
    role: Role


class QuoteItemIn(BaseModel):
    sku: str
    quantity: float


class CalculateIn(BaseModel):
    line: SupplierLine
    customer_name: Optional[str] = None
    contract_months: int = 36
    analyst_hours: float = 0
    exchange_rate: float = 1
    items: List[QuoteItemIn]


class SettingUpdate(BaseModel):
    tax_percent: float
    markup_percent: float
    hour_cost: float


def normalize_header(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def parse_decimal(value) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    txt = str(value).strip()
    if not txt:
        return None
    txt = txt.replace("R$", "").replace("US$", "").replace("$", "").replace(" ", "")
    if "," in txt and "." in txt:
        txt = txt.replace(".", "").replace(",", ".")
    else:
        txt = txt.replace(",", ".")
    try:
        return Decimal(txt)
    except Exception:
        return None


def find_header_and_columns(ws, requested_sku: str, requested_desc: str, requested_cost: str, requested_type: str):
    aliases = {
        "sku": {
            normalize_header(requested_sku),
            "part",
            "part_number",
            "partnumber",
            "pn",
            "sku",
            "item_code",
            "codigo",
            "codigo_item",
            "product_code",
        },
        "desc": {
            normalize_header(requested_desc),
            "desc",
            "description",
            "item_description",
            "product_description",
            "nome",
            "name",
            "produto",
        },
        "cost": {
            normalize_header(requested_cost),
            "price",
            "unit_price",
            "unit_cost",
            "cost",
            "valor",
            "preco",
            "preco_unitario",
            "list_price",
            "msrp",
        },
        "type": {
            normalize_header(requested_type),
            "type",
            "categoria",
            "category",
            "class",
            "item_type",
        },
    }

    best = None
    exact_header_idx = None
    for row_num in range(1, min(120, ws.max_row) + 1):
        row_values = [normalize_header(cell) for cell in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True).__next__()]
        row_map = {h: idx for idx, h in enumerate(row_values) if h}
        if "sku" in row_map:
            exact_header_idx = (row_num, row_map["sku"])
        score = 0
        indices = {}
        for key in ("sku", "desc", "cost"):
            found = next((row_map[a] for a in aliases[key] if a in row_map), None)
            if found is not None:
                score += 1
                indices[key] = found
        type_idx = next((row_map[a] for a in aliases["type"] if a in row_map), None)
        if type_idx is not None:
            indices["type"] = type_idx
        if score >= 3:
            return row_num, indices
        if best is None or score > best[0]:
            best = (score, row_num, indices)

    # Heuristic fallback: infer columns by data patterns when headers are unusual.
    if ws.max_row < 3:
        return (best[1], best[2]) if best and best[0] >= 2 else (None, None)

    probe_start = (best[1] + 1) if best else 2
    probe_end = min(probe_start + 120, ws.max_row)
    max_col = ws.max_column
    col_stats = []

    for col_idx in range(max_col):
        values = []
        numeric_count = 0
        text_count = 0
        sku_like_count = 0

        for row_num in range(probe_start, probe_end + 1):
            row = ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True).__next__()
            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                continue
            s = str(val).strip()
            if not s:
                continue
            values.append(s)
            if parse_decimal(val) is not None:
                numeric_count += 1
            else:
                text_count += 1
            if re.search(r"[A-Za-z]", s) and re.search(r"[0-9]", s):
                sku_like_count += 1

        col_stats.append(
            {
                "idx": col_idx,
                "numeric": numeric_count,
                "text": text_count,
                "sku_like": sku_like_count,
                "samples": values[:5],
                "avg_len": (sum(len(v) for v in values) / len(values)) if values else 0,
            }
        )

    if not col_stats:
        return (best[1], best[2]) if best and best[0] >= 2 else (None, None)

    cost_candidate = max(col_stats, key=lambda c: c["numeric"])

    if exact_header_idx is not None:
        sku_candidate = {"idx": exact_header_idx[1], "sku_like": 999}
    else:
        # SKU tends to be alphanumeric but shorter and with fewer spaces than description.
        sku_candidate = max(col_stats, key=lambda c: (c["sku_like"], -c["avg_len"], c["text"]))

    desc_pool = [c for c in col_stats if c["idx"] not in {sku_candidate["idx"], cost_candidate["idx"]}]
    if not desc_pool:
        desc_pool = col_stats
    # Description tends to be text-heavy and longer.
    desc_candidate = max(desc_pool, key=lambda c: (c["text"], c["avg_len"]))

    inferred = {}
    if sku_candidate["idx"] is not None:
        inferred["sku"] = sku_candidate["idx"]
    if desc_candidate["text"] > 0:
        inferred["desc"] = desc_candidate["idx"]
    if cost_candidate["numeric"] > 0:
        inferred["cost"] = cost_candidate["idx"]

    if len({"sku", "desc", "cost"} & set(inferred.keys())) == 3:
        return (best[1] if best else 1, inferred)

    return (best[1], best[2]) if best and best[0] >= 2 else (None, None)


def db_session():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def bootstrap(db: Session):
    if not db.query(User).count():
        db.add_all([
            User(username="admin", password="admin123", role=Role.admin),
            User(username="vendas", password="vendas123", role=Role.sales),
        ])
    for line in SupplierLine:
        exists = db.query(Setting).filter(Setting.line == line).first()
        if not exists:
            db.add(Setting(line=line))
    db.commit()


def create_token(username: str, role: Role):
    exp = datetime.now(timezone.utc) + timedelta(hours=8)
    return jwt.encode({"sub": username, "role": role.value, "exp": exp}, SECRET, algorithm=ALGO)


def get_current_user(
    cred: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(db_session)
):
    try:
        payload = jwt.decode(cred.credentials, SECRET, algorithms=[ALGO])
    except Exception as exc:
        raise HTTPException(status_code=401, detail="Token inválido") from exc
    user = db.query(User).filter(User.username == payload.get("sub")).first()
    if not user:
        raise HTTPException(status_code=401, detail="Usuário não encontrado")
    return user


def require_admin(user: User = Depends(get_current_user)):
    if user.role != Role.admin:
        raise HTTPException(status_code=403, detail="Apenas admin")
    return user


@app.on_event("startup")
def startup():
    with SessionLocal() as db:
        bootstrap(db)


@app.post("/auth/login", response_model=TokenOut)
def login(body: LoginIn, db: Session = Depends(db_session)):
    user = db.query(User).filter(User.username == body.username, User.password == body.password).first()
    if not user:
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    return TokenOut(access_token=create_token(user.username, user.role), role=user.role)


@app.get("/catalog/{line}")
def catalog(line: SupplierLine, q: str = "", db: Session = Depends(db_session), user: User = Depends(get_current_user)):
    active = db.query(PriceListVersion).filter(PriceListVersion.line == line, PriceListVersion.is_active.is_(True)).first()
    if not active:
        return []
    items = db.query(PriceItem).filter(PriceItem.version_id == active.id)
    if q:
        q_like = f"%{q}%"
        items = items.filter((PriceItem.sku.like(q_like)) | (PriceItem.description.like(q_like)))
    return [{"sku": i.sku, "description": i.description, "unit_cost": float(i.unit_cost), "currency": i.currency, "type": i.type} for i in items.limit(200).all()]


@app.get("/admin/settings/{line}")
def get_settings(line: SupplierLine, db: Session = Depends(db_session), user: User = Depends(get_current_user)):
    st = db.query(Setting).filter(Setting.line == line).first()
    if not st:
        raise HTTPException(status_code=404, detail="Linha sem configuração")
    return {"line": st.line.value, "tax_percent": float(st.tax_percent), "markup_percent": float(st.markup_percent), "hour_cost": float(st.hour_cost)}


@app.put("/admin/settings/{line}")
def update_settings(line: SupplierLine, body: SettingUpdate, db: Session = Depends(db_session), admin: User = Depends(require_admin)):
    st = db.query(Setting).filter(Setting.line == line).first()
    if not st:
        st = Setting(line=line)
        db.add(st)
    st.tax_percent = Decimal(str(body.tax_percent))
    st.markup_percent = Decimal(str(body.markup_percent))
    st.hour_cost = Decimal(str(body.hour_cost))
    db.commit()
    return {"ok": True}


@app.post("/admin/import-price-list")
def import_price_list(
    line: SupplierLine = Form(...),
    supplier_name: str = Form(...),
    sku_column: str = Form("part"),
    desc_column: str = Form("desc"),
    cost_column: str = Form("price"),
    currency: str = Form("BRL"),
    type_column: str = Form("type"),
    file: UploadFile = File(...),
    db: Session = Depends(db_session),
    admin: User = Depends(require_admin),
):
    wb = load_workbook(file.file, data_only=True)

    selected = None
    selected_header_row = None
    selected_idx = None
    best_score = -1

    for ws in wb.worksheets:
        header_row, idx = find_header_and_columns(ws, sku_column, desc_column, cost_column, type_column)
        score = 0 if idx is None else len([k for k in ("sku", "desc", "cost") if k in idx])
        if score > best_score:
            best_score = score
            selected = ws
            selected_header_row = header_row
            selected_idx = idx
        if score == 3:
            break

    if not selected or not selected_idx or any(k not in selected_idx for k in ("sku", "desc", "cost")):
        raise HTTPException(
            status_code=400,
            detail="Nao foi possivel identificar colunas obrigatorias no XLSX (SKU, descricao e custo).",
        )

    db.query(PriceListVersion).filter(PriceListVersion.line == line, PriceListVersion.is_active.is_(True)).update({"is_active": False})

    version = PriceListVersion(line=line, supplier_name=supplier_name, uploaded_by=admin.username, is_active=True)
    db.add(version)
    db.flush()

    imported = 0
    seen = set()
    skipped = 0
    for row in selected.iter_rows(min_row=(selected_header_row or 1) + 1, values_only=True):
        sku_val = row[selected_idx["sku"]] if selected_idx["sku"] < len(row) else None
        desc_val = row[selected_idx["desc"]] if selected_idx["desc"] < len(row) else ""
        cost_val = row[selected_idx["cost"]] if selected_idx["cost"] < len(row) else None
        type_val = row[selected_idx["type"]] if "type" in selected_idx and selected_idx["type"] < len(row) else "default"

        sku = str(sku_val or "").strip()
        if not sku:
            skipped += 1
            continue
        cost = parse_decimal(cost_val)
        if cost is None:
            skipped += 1
            continue
        if sku in seen:
            continue
        seen.add(sku)

        db.add(
            PriceItem(
                version_id=version.id,
                sku=sku,
                description=str(desc_val or "").strip(),
                unit_cost=cost,
                currency=currency.upper(),
                type=str(type_val or "default").strip(),
            )
        )
        imported += 1

    db.commit()
    return {
        "ok": True,
        "version_id": version.id,
        "imported": imported,
        "skipped": skipped,
        "sheet_used": selected.title,
        "header_row": selected_header_row,
    }


@app.post("/pricing/calculate")
def calculate(body: CalculateIn, db: Session = Depends(db_session), user: User = Depends(get_current_user)):
    active = db.query(PriceListVersion).filter(PriceListVersion.line == body.line, PriceListVersion.is_active.is_(True)).first()
    if not active:
        raise HTTPException(status_code=400, detail="Sem price list ativa para a linha")
    st = db.query(Setting).filter(Setting.line == body.line).first()

    subtotal = Decimal("0")
    breakdown = []
    for req_item in body.items:
        item = db.query(PriceItem).filter(PriceItem.version_id == active.id, PriceItem.sku == req_item.sku).first()
        if not item:
            raise HTTPException(status_code=404, detail=f"SKU não encontrado: {req_item.sku}")
        line_total = Decimal(str(req_item.quantity)) * item.unit_cost * Decimal(str(body.exchange_rate))
        subtotal += line_total
        breakdown.append({"sku": item.sku, "description": item.description, "quantity": req_item.quantity, "unit_cost": float(item.unit_cost), "line_total": float(line_total)})

    service_cost = Decimal(str(body.analyst_hours)) * st.hour_cost
    base = subtotal + service_cost
    taxed = base * (Decimal("1") + (st.tax_percent / Decimal("100")))
    monthly = taxed * (Decimal("1") + (st.markup_percent / Decimal("100")))
    total_contract = monthly * Decimal(str(body.contract_months))

    q = Quote(
        line=body.line,
        created_by=user.username,
        customer_name=body.customer_name,
        payload=str(breakdown),
        total_monthly=monthly.quantize(Decimal("0.01")),
        total_contract=total_contract.quantize(Decimal("0.01")),
    )
    db.add(q)
    db.commit()

    return {
        "quote_id": q.id,
        "line": body.line.value,
        "base_cost": float(base),
        "monthly_price": float(q.total_monthly),
        "contract_total": float(q.total_contract),
        "tax_percent": float(st.tax_percent),
        "markup_percent": float(st.markup_percent),
        "breakdown": breakdown,
    }


@app.get("/integrations/ixc/quote/{quote_id}")
def quote_for_ixc(quote_id: int, db: Session = Depends(db_session), user: User = Depends(get_current_user)):
    q = db.query(Quote).filter(Quote.id == quote_id).first()
    if not q:
        raise HTTPException(status_code=404, detail="Cotação não encontrada")
    return {
        "solution": q.line.value,
        "quote_id": q.id,
        "customer_name": q.customer_name,
        "monthly_price": float(q.total_monthly),
        "contract_total": float(q.total_contract),
        "created_at": q.created_at.isoformat(),
    }


@app.get("/health")
def health():
    return {"ok": True}
